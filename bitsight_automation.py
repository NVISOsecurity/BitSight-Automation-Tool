#!/usr/bin/python3

try:    
    from BitSightAPI.client import Session
    from BitSightAPI.companies import Companies
    import json
    from datetime import datetime
    import csv
    from sys import exit
    import openpyxl
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
    import ArgumentsHandler
    import re
    import os
    import sys
    import traceback
    from dateutil.relativedelta import relativedelta
except ModuleNotFoundError as ex:
    file = str(ex).split("'")[-2]
    print(f"[-] Module '{file}' not Found! Please make sure you have installed all the dependencies correctly and all the files are in the right place.")
    exit(1) 
# BitSight Automation tool.

api_key = os.getenv("BITSIGHT_API_KEY")
if api_key == None:
    print("API Key not in environment variables. Add it to your environment variables before executing!\nIf you just added it, close this terminal and open a new one for the environment to refresh...")
    exit(1)

if len(api_key) != 40:
    print("API Key invalid!")
    exit(1)
    
session = Session(api_key)
arghandler = None

def get_entity_rating(guid, entity):
    cmp = Companies(session, guid)
    j = cmp.info('company details')
    
    return j['ratings'][0]['rating']
    
def rating(entity, guid_map, group_map, sort, verbose, group=True):

    if not group:
        score = check_rating(guid_map[entity], entity)
        print(entity + " - " + str(score))
        return entity + " - " + str(score)

    if type(group_map) is dict and list(group_map.keys())[0] == entity:
        ratings = calc_recursive_rating(entity, guid_map, group_map[entity])
        if sort == 'alphabetically': 
            for r in ratings:
                print(r)
            save_rating(ratings, entity)
            return ratings
        else:
            scores = []
            entities = []
            new_ratings = []
            for r in ratings:
                try:
                    scores.append(int(r.split('-')[-1]))
                    entities.append(r.split('-')[0])
                except:
                    continue
            sorted_scores, sorted_entities = zip(*sorted(zip(scores, entities)))
            for i in range(len(sorted_entities)):
                new_ratings.append(sorted_entities[i] + " - " + str(sorted_scores[i]))
                print(sorted_entities[i] + " - " + str(sorted_scores[i]))
            save_rating(new_ratings, entity)
            return new_ratings
    else:
        for g in group_map:
            if type(group_map) is dict and type(group_map[g]) is list:
                for subgroup in group_map[g]:
                    rating(entity, guid_map, subgroup, sort, verbose, True)

def calc_recursive_rating(entity, guid_map, group_map):
    ratings = []
    print(f"Working on {entity}...")
    ratings.append(entity + " - " + str(check_rating(guid_map[entity], entity)))
    
    if type(group_map) is str:
        return ratings
    elif type(group_map) is list:
        for e in group_map:
            if type(e) is str:
                ratings.append(e + " - " + str(check_rating(guid_map[e], e)))
            else:
                temp = calc_recursive_rating(list(e.keys())[0], guid_map, list(e.values())[0])
                for t in temp:
                    ratings.append(t)
    else:
        temp = calc_recursive_rating(list(group_map.keys())[0], guid_map, list(group_map.values()))
        for t in temp:
            ratings.append(t)
    return ratings


def check_rating(guid, entity):
    try:
        if type(guid) is list:
            for g in guid:
                return get_entity_rating(g, entity)
        else:
            if guid == "-":
                print(f"[-] This entity: {entity} does not map with a GUID. It is going to be skipped.")
                return

            return get_entity_rating(guid, entity)

    except KeyError:
            print("[-] Cannot retrieve information for this entity. Perhaps the subscription got revoked. Skipping...")
            return None

def save_rating(data, entity):
    with open(f"{datetime.now().date()}_bitsight_rating_{entity}.txt", 'w') as handle:
        if type(data) is list:
            for row in data:
                handle.write(row + "\n")
        else:
            handle.write(data)
    print(f"[+] Data saved to: {datetime.now().date()}_bitsight_rating_{entity}.txt")

def retrieve_findings(guid):
    cmp = Companies(session, guid)
    j = cmp.info('findings', {'limit':100000})
    return j

def convert_json_to_csv(entity, json_data, severity, verbose):
    try:
        headers = list(json_data['results'][0].keys())
        headers.append('remediation_tips')
        count = 0
        with open(f'bitsight_{severity}_findings_{entity}_{str(datetime.now().date())}.csv', 'w', newline='', encoding='utf-8') as handle:
            csv_writer = csv.writer(handle)
            csv_writer.writerow(headers)

            # Filtering
            for finding in json_data['results']:
                if severity == 'Low':
                    if not finding['severity_category'] == 'minor':
                        continue
                if severity == 'Medium':
                    if not finding['severity_category'] == 'moderate':
                        continue
                if severity == 'High':
                    if not finding['severity_category'] == 'material':
                        continue
                if severity == 'Critical':
                    if not finding['severity_category'] == 'severe':
                        continue
                if severity == 'Critical/High':
                    if not (finding['severity_category'] == 'material' or finding['severity_category'] == 'severe'):
                        continue

                # Saving
                row = []
                remediation_tips = []
                for header in headers:
                    if header == 'assets':
                        assetlist = []
                        for asset in finding[header]:
                            assetlist.append(asset['asset'])
                        row.append(', '.join(assetlist))
                    elif header == 'details': 
                        vuln_details = []
                        for vuln in finding[header]['remediations']:
                            vuln_details.append(vuln['help_text'])
                            remediation_tips.append(vuln['remediation_tip'])
                        row.append(' | '.join(vuln_details))
                    elif header == 'remediation_tips':
                        pass
                    else:
                        row.append(finding[header])
                row.append(remediation_tips)
                csv_writer.writerow(row)
                count += 1
        if verbose:
            print("Findings count: " + str(count))
        print(f'[+] Data saved to bitsight_{severity}_findings_{entity}_{str(datetime.now().date())}.csv')
    except IndexError as ex:
        print("It appears as there are no findings in " + entity + " or there is something wrong with the API. Please validate the old fashioned way using your browser.\nMore Details: "+ str(ex))
        print("It might be the case you do not have a 'Total Risk Monitoring' subscription. The 'Risk Monitoring' subscription is unable to work with the API for this operation")
        print()
        print('Response: ' + str(json_data))
    except KeyError as ex:
        print(f"'Total Risk Monitoring' subscription is required for this operation and {entity} does not have one.")
        
def retrieve_assets(guid, entity, verbose):
    try:
        cmp = Companies(session, guid)
        ips = cmp.info('ip by country', {'limit':10000})

        for ip in ips['ipv4']:
            print(ip + " - " + str(ips['ipv4'][ip]))
    except KeyError as ex:
        print("Cannot retrieve total count of IPs for " + entity + "\nA 'Total Risk Monitoring' subscription is required for this operation.\nMore Details: " + str(ips))
    try:
        assets = cmp.info('asset risk matrix')
        print('*********** Asset List ************')
        with open(f'bitsight_asset_list_{entity}_{datetime.now().date()}.csv', 'w', newline='') as handle:
            csv_writer = csv.writer(handle)
            csv_writer.writerow(['Assets'])
            for asset in assets['assets']:
                if verbose:
                    print(asset['asset'])
                csv_writer.writerow([asset['asset']])
        print(f'[+] Asset List saved to: bitsight_asset_list_{entity}_{datetime.now().date()}.csv')
    except KeyError as ex:
        print("Cannot retrieve asset list for " + entity + "\nA 'Total Risk Monitoring' subscription is required for this operation.\nMore Details: " + str(assets))

def reverse_lookup(guid_map, group_map, search):
    cmp = Companies(session, guid_map[list(group_map.keys())[0]])
    asset_list = cmp.info('assets', {'limit':30000})
    wildcard_name = False
    iprange = False
    if '*' in search:
        wildcard_name = True
        search = search.replace('*','')
    if '/' in search:
        iprange = search.split('/')[-1]
        if int(iprange) >= 24:
            search = search.split('/')[0][:-2]
        elif int(iprange) >= 16:
            search = search.split('/')[0][:-4]
        elif int(iprange) >= 7:
            search = search.split('/')[0][:-6]
        else:
            search = ''
    for asset in asset_list['results']:
        try:
            asset_name = asset['asset']
            ips = asset['ip_addresses']
            guid = asset['origin_subsidiary']['guid']
            if len(ips) > 0:
                for ip in ips:
                    if (search == ip and not iprange) or (search in ip and iprange):
                        print(f"{asset_name} - {ips}: Found in: {list(guid_map.keys())[list(guid_map.values()).index(guid)]}")
                        break
            if (search == asset_name and not wildcard_name) or (search in asset_name and wildcard_name):
                print(f"{asset_name} - {ips}: Found in: {list(guid_map.keys())[list(guid_map.values()).index(guid)]}")
        except TypeError:
            pass
        except ValueError:
            pass


def init():

    print("Files are empty, initializing tool... Gathering information for your convenience... This may take a while...\n")
    print("If this is your first execution of this tool, you can consider it normal...")

    cmp = Companies(session)
    result = cmp.info()
    with open('companies.csv', 'w', newline='') as handle:
        csv_writer = csv.writer(handle)
        csv_writer.writerow(['GUID', ' Subsidiary Name'])
        for company in result['companies']:
            guid = company['guid']
            name = company['name']
            
            csv_writer.writerow([guid, name])
    print("[+] File saved at companies.csv")
    print("Please review the file and manually add the entries in the json files.")
    exit()


def update(guid_map, group_map):
    # Some entities have multiple guids. (More than one subcompanies are assigned to them.)
    guid_clear = []
    for value in guid_map.values():
        if type(value) is list:
            for item in value:
                guid_clear.append(item)
        else:
            guid_clear.append(value)


    cmp = Companies(session)
    result = cmp.info()
    for company in result['companies']:
        guid = company['guid']
        name = company['name']
        if guid not in guid_clear:
            print(name + " - " + guid + " not found in our configuration\n")
            decision = input("Would you like to include it (Y/N)? ")
            if decision.lower() == 'y':
                entity_code = input("What is the name of this entity? ")
                while True:
                    group = input(f"Under which group should this entity fall under ({arghandler.groups}) ? ")
                    if group not in arghandler.groups:
                        print("Invalid group, please try again")
                    else:
                        break
                print("Adding " + name + " with guid {" + guid + "} as " + entity_code + " in " + group + "\n\n")
                guid_map[entity_code] = guid
                if entity_code not in arghandler.groups:
                    add_to_right_path(group_map, entity_code, group)
                save_configuration(guid_map, group_map)
                print()
                print()

def add_to_right_path(group_map, entity_code, group):
    for key, value in group_map.items():
        if key == group:
            for item in group_map[key]:
                if item == entity_code:
                    return
            group_map[key].append(entity_code)
        if type(value) is list:
            for item in value:
                if type(item) is dict:
                    add_to_right_path(item, entity_code, group)


def save_configuration(guid_map, group_map):
    with open('guid_mapper.json','w') as handle1, open('group_mapper.json','w') as handle2:
        json.dump(guid_map, handle1, indent="", ensure_ascii=True)
        json.dump(group_map, handle2, indent="", ensure_ascii=True)
        print("Configuration Updated")
    

def load_configs(): 
    try:
        with open("group_mapper.json") as handle:
            group_map = json.loads(handle.read())

        with open("guid_mapper.json") as handle:
            guid_map = json.loads(handle.read())

        if len(guid_map) == 0 and len(group_map) == 0:
            init()

        return group_map, guid_map
    except FileNotFoundError as ex:
        print(f"File not found: {str(ex).split(':')[1]}. Please copy the {str(ex).split(':')[1]} to the same directory as this tool and try again.")
        exit(1)

def list_config(guid_map):
    print("Listing Configuration...")
    cmp = Companies(session)
    result = cmp.info()
    list_dict = {}
    for company in result['companies']:
        guid = company['guid']
        name = company['name']
        list_dict[guid] = name
    for key in guid_map.keys():
        if type(guid_map[key]) is list:
            guid = guid_map[key][0]
        else:
            guid = guid_map[key]
        try:
            print(key + " - " + list_dict[guid])
        except:
            print("No mapping for: " + key)

def historical(entity, guid, months_back):
    try:
        if type(guid) is list:
            guid = guid[0]
        cmp = Companies(session, guid)
        j = cmp.info('company details')
        all_months = {}
        for rating in j['ratings']:
                score = rating['rating']
                date = rating['rating_date']
                all_months[date] = score
        now = datetime.now().date()
        months_remaining = int(months_back)
        data = {}
        data[entity] = []
        entity_rating = {}
        errors = 0
        while months_remaining >= 0:
            if str(now - relativedelta(months=months_remaining, days=5)) in all_months.keys():
                entity_rating[str(now - relativedelta(months=months_remaining, days=5))] = str(all_months[str(now - relativedelta(months=months_remaining, days=5))])
            else:
                errors += 1
            entity_rating[str(now - relativedelta(months=months_remaining, days=5))] = str(all_months[str(now - relativedelta(months=months_remaining, days=5))])
            months_remaining -= 1
        if errors > 4:
            print("Something is not working properly. Have a look at the API response to validate how many months back you can retrieve data for.\n\nTo do so, you can use print(all_months) right before the while loop above...")
            exit(1)
        data[entity].append(entity_rating)
        return data
    except KeyError as ex:
        print(entity + "'s data cannot be retrieved.\nMost Likely there is no active subscription for this entity.\nError: " + str(ex) + "... Skipping entity.")
        return

            
def historical_group(entity, guid_map, months, group_map, ratings_list):
    for key,value in group_map.items():
        if key == entity:
            print(f"Working on {entity}..." )   
            ratings_list.append(historical(entity, guid_map[entity], months))
            if type(group_map[entity]) is list:
                for item in group_map[entity]:
                    if type(item) is not dict:
                        ratings_list.append(historical(item, guid_map[item], months))
                    elif type(item) is dict:
                        historical_group(list(item.keys())[0], guid_map, months, item, ratings_list)
            return
    
        if type(value) is list:
            for item in value:
                if type(item) is str:
                    continue
                historical_group(entity, guid_map, months, item, ratings_list)
            

def save_historical_data(entity, data, months):
    try:
        workbook = openpyxl.Workbook()
        # Create Legend
        sheet = workbook.create_sheet('Legend')
        sheet.cell(row=1, column=1).value = "Score < 650"
        sheet.cell(row=1, column=1).fill = PatternFill(start_color='AF4353', fill_type='solid')
        sheet.cell(row=2, column=1).value = 'Score < 740'
        sheet.cell(row=2, column=1).fill = PatternFill(start_color='E8A753', fill_type='solid')
        sheet.cell(row=3, column=1).value = 'Score >= 740'
        sheet.cell(row=3, column=1).fill = PatternFill(start_color='2C4D82', fill_type='solid')
        sheet.cell(row=4, column=1).value = "Note: An empty row indicates the results could not be fetched."
        # Store Data
        sheet = workbook.active
        sheet.title = 'Historical Scores'
        headers = ['Subsidiary/Date']
        try:
            headers += list(list(data[0].values())[0][0].keys())
        except:
            headers += list(list(data[1].values())[0][0].keys())
        rows = len(data) +1     
        columns = int(months)+2
        sheet.append(headers)
        for entry in data:
            row = []
            try:
                row.append(list(entry.keys())[0])
                for key in list(entry.values())[0][0].keys():
                    row.append(list(entry.values())[0][0][key])
            except AttributeError:
                pass

            sheet.append(row)
        beautify_excel_structure(sheet, rows, columns)
        workbook.save(f'{datetime.now().date()}_{entity}_bitsight_historical_ratings_{months}_months.xlsx')
        print(f"[+] Data saved to {datetime.now().date()}_{entity}_bitsight_historical_ratings_{months}_months.xlsx")
    except AttributeError as ex:
        print("No data has been deivered back. Error: " + str(ex))
        exit(1)      

def beautify_excel_structure(sheet, rows, columns):
    border = Border(top=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
    ### Header section
    for i in range(1, columns+1):
        sheet.cell(row=1,column=i).fill = PatternFill(start_color='cccccc', fill_type='solid')
        sheet.cell(row=1,column=i).border = border
    for i in range(1, rows+1):
        sheet.cell(row=i,column=1).fill = PatternFill(start_color='cccccc', fill_type='solid')
        sheet.cell(row=i,column=1).border = border
    ### Data section
    for i in range(1, rows+1):
        # Replace with your groups below...
        if sheet.cell(row=i, column=1).value in arghandler.groups:
            for j in range(1, columns+1):
                sheet.cell(row=i, column=j).font = Font(bold=True)
    for i in range(2, rows+1):
        for j in range(2,columns+1):
            try:
                if int(sheet.cell(row=i,column=j).value) < 650:
                    color = 'AF4353'
                elif int(sheet.cell(row=i,column=j).value) < 740:
                    color = 'E8A753'
                else:
                    color = '2C4D82'
                sheet.cell(row=i, column=j).fill = PatternFill(start_color=color, fill_type='solid')
            except TypeError:
                pass
def main():
    global arghandler
    arghandler = ArgumentsHandler.ArgumentHandler()

    group_map, guid_map = load_configs()

    if arghandler.operation == 'rating':
        if arghandler.entity:
            save_rating(rating(arghandler.entity, guid_map, group_map, arghandler.sort, arghandler.verbose, False), arghandler.entity)
        else:
            print("[*] This may take a moment. Grab a coffee")
            rating(arghandler.group, guid_map, group_map, arghandler.sort, arghandler.verbose)
    elif arghandler.operation == 'findings':
        if not arghandler.entity:
            print("This operation can only work with the entity argument.")
            exit(1)
        if arghandler.entity in arghandler.groups:
            print("This operation can only work with specified entities, not groups. Please try again")
            exit(1)
        if arghandler.entity not in guid_map.keys():
            print('This entity is not in the available entities. Please try another one.')
            exit(1)
        convert_json_to_csv(arghandler.entity, retrieve_findings(guid_map[arghandler.entity]), arghandler.severity, arghandler.verbose)
    elif arghandler.operation == 'assets':
        # Assets
        if not arghandler.entity:
            print("This operation can only work with the entity argument.")
            exit(1)
        if arghandler.entity in arghandler.groups:
            print("This operation can only work with specified entities, not groups. Please try again")
            exit(1)
        retrieve_assets(guid_map[arghandler.entity], arghandler.entity, arghandler.verbose)    
    elif arghandler.operation == 'reverse_lookup':
        reverse_lookup(guid_map, group_map, arghandler.search)
    elif arghandler.operation == 'update':
        update(guid_map, group_map)
        print("Update completed!")
    elif arghandler.operation == 'historical':
        print("Grab a coffee, this will take a while...")
        ratings_list = []
        if arghandler.entity:
            ratings_list.append(historical(arghandler.entity, guid_map[arghandler.entity], arghandler.months))
            save_historical_data(arghandler.entity, ratings_list, arghandler.months)   
        else:
            historical_group(arghandler.group, guid_map, arghandler.months, group_map, ratings_list)
            save_historical_data(arghandler.group, ratings_list, arghandler.months)
    elif arghandler.operation == 'list':
        list_config(guid_map)



if __name__ == '__main__':
    try:
        main()
    except Exception as ex:
        if 'socket' in str(ex) or 'connection' in str(ex):
            print("The tool was unable to connect to BitSight's server. If the tool is behind a Firewall it might be blocked or you may not have an internet connection.")
            exit(1)
        else:
            print("Please verify that the API key you are using is the correct...")
            print("Some other issue has been noted. Please open a GitHub Issue with the traceback below.")
            print("Share this traceback...")
            print("=" * 30)
            traceback.print_exc(file=sys.stdout)
            print("=" * 30)
            exit(1)
